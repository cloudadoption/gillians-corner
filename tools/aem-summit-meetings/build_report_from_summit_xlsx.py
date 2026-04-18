# -*- coding: utf-8 -*-
"""Rebuild BUILTIN_PER_PERSON, BUILTIN_DATA, REGION_MAP (merge) from Summit CBC xlsx for report.html."""
from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

REPORT_PATH = Path(__file__).resolve().parent / "report.html"
XLSX_PATH = Path(
    r"c:\Users\gkrause\OneDrive - Adobe\Documents\Adobe\SUMMIT26\Summit CBC Meeting Report 04_17 730am PT.xlsx"
)
SHEET = "04_17 730am"

GEO_MAP = {
    "US/Canada": "US/Canada",
    "EMEA": "EMEA",
    "APAC": "APAC",
    "Japan": "APAC",
    "Latin America/Mexico": "Latin America",
}


def norm_spaces(s: str) -> str:
    return " ".join((s or "").split())


def extract_attendees_block(html: str) -> str:
    m = re.search(r"const ATTENDEES=\[([\s\S]*?)\];", html)
    if not m:
        raise SystemExit("ATTENDEES block not found")
    return m.group(1)


def parse_attendee_names(att_block: str) -> list[str]:
    return re.findall(r"name:'([^']*)'", att_block)


def parse_name_aliases(html: str) -> dict[str, str]:
    m = re.search(r"const NAME_ALIASES=\{([^}]*)\};", html, re.DOTALL)
    if not m:
        return {}
    inner = m.group(1)
    out: dict[str, str] = {}
    for km in re.finditer(r"'([^']+)':'([^']+)'", inner):
        out[km.group(1)] = km.group(2)
    return out


def parse_region_map(html: str) -> dict[str, str]:
    m = re.search(r"let REGION_MAP=(\{[^;]+\});", html)
    if not m:
        raise SystemExit("REGION_MAP not found")
    return json.loads(m.group(1))


def canonical_name(raw: str, aliases: dict[str, str]) -> str:
    s = norm_spaces(raw)
    return aliases.get(s, s)


def region_for_geo(raw) -> str:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return "Unknown"
    key = str(raw).strip()
    return GEO_MAP.get(key, "Unknown")


def compute_meeting_map(
    per_person: dict[str, list[dict]], attendee_names: list[str], region_map: dict[str, str]
) -> dict[str, dict]:
    code_to_attendees: dict[str, set[str]] = defaultdict(set)
    for name in attendee_names:
        for mtg in per_person.get(name, []):
            code_to_attendees[mtg["code"]].add(name)
    multi_codes = {c for c, s in code_to_attendees.items() if len(s) > 1}

    out: dict[str, dict] = {}
    for name in attendee_names:
        agg: dict[str, dict] = defaultdict(
            lambda: {
                "codes": set(),
                "multiCodes": set(),
                "accounts": set(),
                "multiAccounts": set(),
            }
        )
        all_codes: set[str] = set()
        all_multi: set[str] = set()

        for m in per_person.get(name, []):
            region = region_map.get(m["account"], "Unknown")
            if region == "Unknown":
                continue
            is_multi = m["code"] in multi_codes
            d = agg[region]
            d["codes"].add(m["code"])
            all_codes.add(m["code"])
            if is_multi:
                d["multiCodes"].add(m["code"])
                all_multi.add(m["code"])
            d["accounts"].add(m["account"])
            if is_multi:
                d["multiAccounts"].add(m["account"])

        by_region: dict = {}
        multi_by_region: dict = {}
        for r, d in agg.items():
            by_region[r] = {
                "count": len(d["codes"]),
                "multi_count": len(d["multiCodes"]),
                "accounts": sorted(d["accounts"]),
            }
            if d["multiCodes"]:
                multi_by_region[r] = {
                    "count": len(d["multiCodes"]),
                    "accounts": sorted(d["multiAccounts"]),
                }
        out[name] = {
            "total": len(all_codes),
            "multi_total": len(all_multi),
            "by_region": by_region,
            "multi_by_region": multi_by_region,
        }
    return out


def load_meetings_from_xlsx(
    aliases: dict[str, str],
    attendee_set: set[str],
    region_map_existing: dict[str, str],
) -> tuple[dict[str, list[dict]], dict[str, str]]:
    """Returns per_person meetings (only keys in attendee_set) and account->region for merge."""
    wb = load_workbook(str(XLSX_PATH), read_only=True, data_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    by_code: dict[str, list] = defaultdict(list)
    for row in rows:
        status = row[2]
        if status != "Approved":
            continue
        code = row[0]
        if not code:
            continue
        by_code[code].append(row)

    account_region_updates: dict[str, str] = {}
    per_person: dict[str, list[dict]] = defaultdict(list)

    for code, group in by_code.items():
        first = group[0]
        title = (first[4] or "").strip() or "(no title)"
        account = (first[5] or "").strip() or "Not Listed"
        geo_raw = first[28]
        mapped = region_for_geo(geo_raw)
        if mapped == "Unknown" and account in region_map_existing:
            mapped = region_map_existing[account]
        if account and mapped != "Unknown":
            account_region_updates[account] = mapped

        all_names: set[str] = set()
        summit_names: set[str] = set()
        for row in group:
            fn = row[10]
            if not fn:
                continue
            company = (row[11] or "").strip()
            canon = canonical_name(str(fn), aliases)
            all_names.add(canon)
            ee = row[14]
            is_yes = str(ee).strip().lower() == "yes" if ee is not None else False
            if company == "Adobe" and is_yes:
                summit_names.add(canon)

        all_sorted = sorted(all_names, key=lambda s: s.lower())
        summit_sorted = sorted(summit_names, key=lambda s: s.lower())

        adobe_in_meeting: set[str] = set()
        for row in group:
            fn = row[10]
            if not fn:
                continue
            company = (row[11] or "").strip()
            if company != "Adobe":
                continue
            canon = canonical_name(str(fn), aliases)
            adobe_in_meeting.add(canon)

        mtg = {
            "code": str(code).strip(),
            "title": title,
            "account": account,
            "all_attendees": all_sorted,
            "summit_attendees": summit_sorted,
            "region": mapped,
        }
        for name in adobe_in_meeting:
            if name not in attendee_set:
                continue
            if any(m["code"] == mtg["code"] for m in per_person[name]):
                continue
            per_person[name].append(dict(mtg))

    for name in attendee_set:
        per_person.setdefault(name, [])
        per_person[name].sort(key=lambda m: (m["code"], m["title"]))

    return dict(per_person), account_region_updates


def main() -> None:
    global html_text
    html_text = REPORT_PATH.read_text(encoding="utf-8")
    att_block = extract_attendees_block(html_text)
    attendee_names = parse_attendee_names(att_block)
    attendee_set = set(attendee_names)
    aliases = parse_name_aliases(html_text)
    region_map = parse_region_map(html_text)

    per_person, acc_updates = load_meetings_from_xlsx(
        aliases, attendee_set, region_map
    )
    for acc, reg in acc_updates.items():
        if reg != "Unknown":
            region_map.setdefault(acc, reg)

    builtin_data = compute_meeting_map(per_person, attendee_names, region_map)

    # Order BUILTIN_PER_PERSON keys by ATTENDEES order (include zero-meeting names)
    ordered = {n: per_person.get(n, []) for n in attendee_names}

    per_json = json.dumps(ordered, ensure_ascii=False, separators=(",", ":"))
    data_json = json.dumps(builtin_data, ensure_ascii=False, separators=(",", ":"))
    map_json = json.dumps(region_map, ensure_ascii=False, separators=(", ", ": "))

    out_path = Path(__file__).resolve().parent / "_report_rebuilt_snippets.txt"
    out_path.write_text(
        "===BUILTIN_PER_PERSON===\n"
        + per_json
        + "\n===BUILTIN_DATA===\n"
        + data_json
        + "\n===REGION_MAP===\n"
        + map_json
        + "\n",
        encoding="utf-8",
    )
    print("Wrote", out_path)
    print("Meeting counts:", {n: len(ordered[n]) for n in attendee_names_active})

    # --- Patch report.html ---
    h = html_text
    # BUILTIN_PER_PERSON (one line before const ATTENDEES)
    a = h.find("const BUILTIN_PER_PERSON = ")
    b = h.find("\nconst ATTENDEES=[", a)
    if a < 0 or b < 0:
        raise SystemExit("Could not locate BUILTIN_PER_PERSON block")
    h = h[:a] + "const BUILTIN_PER_PERSON = " + per_json + ";" + h[b:]

    # BUILTIN_DATA
    c = h.find("const BUILTIN_DATA=")
    d = h.find("\nfunction buildDisplayData", c)
    if c < 0 or d < 0:
        raise SystemExit("Could not locate BUILTIN_DATA block")
    h = h[:c] + "const BUILTIN_DATA=" + data_json + ";" + h[d:]

    # REGION_MAP
    rm0 = h.find("let REGION_MAP=")
    rm1 = h.find(";", rm0)
    if rm0 < 0 or rm1 < 0:
        raise SystemExit("Could not locate REGION_MAP")
    h = h[:rm0] + "let REGION_MAP=" + map_json + ";" + h[rm1 + 1 :]

    # Ron Nagy → Engineering
    h = h.replace(
        "{name:'Ron Nagy',title:'Sr. Evangelist',group:'PMM'}",
        "{name:'Ron Nagy',title:'Sr. Evangelist',group:'Engineering'}",
    )

    # Heading / admin copy (spreadsheet timestamp)
    h = h.replace(
        "<title>Summit 2026 — Meeting Dashboard (4/14 data)</title>",
        "<title>Summit 2026 — Meeting Dashboard (4/17 7:30am PT)</title>",
    )
    h = h.replace(
        "<p>Product Experts attending Summit 2026 — built-in data refreshed from the 4/14 EF export (Meetings by employees 4_14 10pm). Same embedded data as index.html; hard-refresh if this label is missing.</p>",
        "<p>Product Experts attending Summit 2026 — built-in data refreshed from <strong>Summit CBC Meeting Report 04_17 730am PT.xlsx</strong> (tab <code>04_17 730am</code>, approved CBC meetings). Adobe name aliases applied (e.g. Sean Steiner → Sean Steimer). Hard-refresh if this label is stale.</p>",
    )
    h = h.replace(
        '<span class="admin-data-text">Latest data: <strong id="dataBannerDate">14 Apr 2026</strong></span>',
        '<span class="admin-data-text">Latest data: <strong id="dataBannerDate">17 Apr 2026 · 7:30am PT</strong></span>',
    )
    h = h.replace(
        '<span class="admin-data-file" id="dataBannerFile">Meeting by Employee 4_14.xlsx · Meetings by employees 4_14 10pm</span>',
        '<span class="admin-data-file" id="dataBannerFile">Summit CBC Meeting Report 04_17 730am PT.xlsx · 04_17 730am</span>',
    )
    h = h.replace(
        '<span id="dataSourceText">Showing built-in data (Meeting by Employee 4_14.xlsx, tab &quot;Meetings by employees 4_14 10pm&quot;, 4/14/2026) — upload a new file above to refresh</span>',
        '<span id="dataSourceText">Showing built-in data (Summit CBC Meeting Report 04_17 730am PT.xlsx, tab &quot;04_17 730am&quot;, 17 Apr 2026 7:30am PT export) — upload a new file above to refresh</span>',
    )

    REPORT_PATH.write_text(h, encoding="utf-8")
    print("Updated", REPORT_PATH)


if __name__ == "__main__":
    main()
